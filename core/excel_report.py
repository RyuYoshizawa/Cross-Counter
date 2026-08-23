"""
excel_report.py
集計表タブの実行結果（core/cross_execute.run_cross_planの戻り値）を、ユーザー提示の
見本Excel（エクセル出力指示.xlsx）どおりの4シート構成・見た目で一括Excel出力する
（SPEC 5.5節、ユーザーとの合意事項 2026-08-22）。
- 「クロス％実数別表」: 2wayクロスを％表・実数表に分けて並べる（現行のアプリ内表示と同じ形式）
- 「旧クロス％実数別表」: 同じ2wayクロスを、旧フォーマット（全体行/列を先頭に置く）で並べる
  ——現行フォーマットに変更する前の見た目に揃えたい既存のレポートテンプレート向け
- 「GT」: 単純集計（GT行）
- 「クロス％＆実数表」: 2wayクロスを、属性カテゴリごとに実数行→％行を上下に並べる1つの表にする

集計表タブ選択中の【集計表形式】設定には関わらず、4シートとも常にまとめて作る
（1回のボタンで全形式を出力してほしいというユーザーの要望）。トリプルクロスはユーザーの
指定した4カテゴリに含まれないため対象外。

**見た目（罫線・塗り・フォント）は見本Excelのセルスタイルを実測して再現**（2026-08-22、
「数値配置は正しいがレイアウトデザインが無い」という指摘を受けて追加）: 表の格子部分には
細罫線、見出し・カテゴリラベルは中央寄せ＋折り返し、フォントは游ゴシック（本文9pt/タイトル
太字10pt）、「全体」行・「全体」列には薄い塗り（見本のセルを実測して確認した配色）を
クロス状に付ける。見本のスタイルはopenpyxlで直接セルを読み込んで確認したもの——
値だけでなくfont/fill/border/alignment/number_format/列幅も実測した。
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

_TOTAL_LABEL = '全体'
_PCT_FORMAT = '0.0%'

_FONT_NAME = '游ゴシック'
_TITLE_FONT = Font(name=_FONT_NAME, size=10, bold=True)
_SUBTITLE_FONT = Font(name=_FONT_NAME, size=10, bold=True)
_BODY_FONT = Font(name=_FONT_NAME, size=9)
_COMMENTARY_FONT = Font(name=_FONT_NAME, size=9)

_THIN = Side(style='thin')
_GRID_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

_TOTAL_FILL = PatternFill(fill_type='solid', fgColor='CCFFFF')

_LABEL_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_AXIS_ALIGN = Alignment(horizontal='right', vertical='center')
_VALUE_ALIGN = Alignment(vertical='center', wrap_text=True)
_COMMENTARY_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=False)

_OPTION_COL_WIDTH = 9.5
_MAX_OPTION_COLS = 20  # 列幅を事前に整えておく範囲（これを超える選択肢数でも値自体は書き込まれる）


SORT_DESC = '全体の出現率が高い方から降順'
SORT_DEFAULT = '対象設問のデフォルト選択肢順'


def build_report_workbook(results: list[dict], *, list_cross_groups: list[dict] | None = None,
                           list_cross_sort_order: str = SORT_DESC) -> Workbook:
    """
    results: core.cross_execute.run_cross_planの戻り値（results側、issuesは含まない）。
    is_gt=TrueのものはGTシートへ、Falseのものは残り3種のクロスシートへ、それぞれ元の順序で
    複数ブロックとして縦に並べる。
    list_cross_groups: core.cross_execute.run_list_crossの戻り値（groups側）——指定があれば
    「一覧型クロス集計表」シートを追加する（SPEC 5.3.3、Excel出力専用の一覧型クロス集計指定表）。
    """
    gt_results = [r for r in results if r.get('is_gt')]
    cross_results = [r for r in results if not r.get('is_gt')]

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('クロス％実数別表')
    _build_cross_split_sheet(ws, cross_results)

    ws = wb.create_sheet('旧クロス％実数別表')
    _build_cross_split_old_sheet(ws, cross_results)

    ws = wb.create_sheet('GT')
    _build_gt_sheet(ws, gt_results)

    ws = wb.create_sheet('クロス％＆実数表')
    _build_cross_combined_sheet(ws, cross_results)

    if list_cross_groups is not None:
        ws = wb.create_sheet('一覧型クロス集計表')
        _build_list_cross_sheet(ws, list_cross_groups, list_cross_sort_order)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _set(ws: Worksheet, row: int, col: int, value, *, font: Font = _BODY_FONT,
         number_format: str | None = None, align: Alignment | None = None,
         border: Border | None = None, fill: PatternFill | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if number_format is not None:
        cell.number_format = number_format
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill
    return cell


def _grid(ws: Worksheet, row: int, col: int, value, *, number_format: str | None = None,
          align: Alignment = _VALUE_ALIGN, is_total: bool = False):
    """表の格子部分（見出し・カテゴリラベル・データ値）を1セル書く共通ヘルパー"""
    return _set(ws, row, col, value, number_format=number_format, align=align,
                border=_GRID_BORDER, fill=_TOTAL_FILL if is_total else None)


def _merge_if_needed(ws: Worksheet, row1: int, row2: int, col: int) -> None:
    if row2 > row1:
        ws.merge_cells(start_row=row1, start_column=col, end_row=row2, end_column=col)


def _set_common_column_widths(ws: Worksheet, *, label_col_width: float = 17.0) -> None:
    ws.column_dimensions['A'].width = 5.5
    ws.column_dimensions['B'].width = 4.0
    ws.column_dimensions['C'].width = 6.0
    ws.column_dimensions['D'].width = label_col_width
    for i in range(5, 5 + _MAX_OPTION_COLS):
        ws.column_dimensions[_col_letter(i)].width = _OPTION_COL_WIDTH


def _col_letter(col: int) -> str:
    letters = ''
    while col > 0:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# ---------------------------------------------------------------- GT

def _build_gt_sheet(ws: Worksheet, gt_results: list[dict]) -> None:
    ws.column_dimensions['A'].width = 5.5
    ws.column_dimensions['B'].width = 4.5
    ws.column_dimensions['C'].width = 24.0
    ws.column_dimensions['D'].width = 9.5
    ws.column_dimensions['E'].width = 9.5
    row = 1
    if not gt_results:
        _set(ws, row, 1, '対象のGT集計がありません。')
        return
    for r in gt_results:
        row = _write_gt_block(ws, row, r)


def _write_gt_block(ws: Worksheet, row: int, r: dict) -> int:
    _set(ws, row, 1, f"{r['target_label']}（GT）", font=_TITLE_FONT)
    row += 2

    _grid(ws, row, 4, '％', align=_LABEL_ALIGN)
    _grid(ws, row, 5, 'ｎ', align=_LABEL_ALIGN)
    row += 1

    table = r['table']
    base = r['base']
    for i, opt_row in enumerate(table.to_dict('records'), 1):
        _grid(ws, row, 2, i, align=_LABEL_ALIGN)
        _grid(ws, row, 3, opt_row['選択肢'], align=_LABEL_ALIGN)
        _grid(ws, row, 4, round(opt_row['%'] / 100, 4), number_format=_PCT_FORMAT)
        _grid(ws, row, 5, int(opt_row['度数']))
        row += 1

    _grid(ws, row, 3, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    _grid(ws, row, 4, 1.0, number_format=_PCT_FORMAT, is_total=True)
    _grid(ws, row, 5, int(base), is_total=True)
    row += 2
    return row


# ---------------------------------------------------------------- クロス％実数別表（現行フォーマット）

def _build_cross_split_sheet(ws: Worksheet, cross_results: list[dict]) -> None:
    _set_common_column_widths(ws)
    row = 1
    if not cross_results:
        _set(ws, row, 1, '対象のクロス集計がありません。')
        return
    for r in cross_results:
        row = _write_cross_split_block(ws, row, r)


def _write_cross_split_block(ws: Worksheet, row: int, r: dict) -> int:
    attr_label, target_label = r['attr_label'], r['target_label']
    pct_df, n_df, base = r['pct'], r['n'], r['base']
    target_cols = [c for c in pct_df.columns if c != _TOTAL_LABEL]
    attr_rows = [i for i in pct_df.index if i != _TOTAL_LABEL]
    n_opts = len(target_cols)
    first_col = 5  # E列
    total_col = first_col + n_opts

    _set(ws, row, 1, f'{attr_label} × {target_label}', font=_TITLE_FONT)
    row += 2

    # ％セクション
    _set(ws, row, 2, f' {target_label}/{attr_label}(％)', font=_SUBTITLE_FONT)
    row += 1
    row = _write_target_group_header(ws, row, target_label, n_opts, first_col)
    _write_index_and_pct_labels(ws, row, target_cols, first_col, total_col, index_label='%', total_label=_TOTAL_LABEL)
    row += 2
    header_row = row
    for cat in [*attr_rows, _TOTAL_LABEL]:
        is_total = cat == _TOTAL_LABEL
        n_val = base.get(cat, 0)
        label = f'{cat}\n(n={n_val})'
        _grid(ws, row, 4, label, align=_LABEL_ALIGN, is_total=is_total)
        for i, col in enumerate(target_cols):
            _grid(ws, row, first_col + i, round(pct_df.loc[cat, col] / 100, 4),
                  number_format=_PCT_FORMAT, is_total=is_total)
        _grid(ws, row, total_col, round(pct_df.loc[cat, _TOTAL_LABEL] / 100, 4),
              number_format=_PCT_FORMAT, is_total=True)
        row += 1
    _write_attr_axis_label(ws, header_row, row - 1, attr_label)
    row += 1

    # 実数セクション
    _set(ws, row, 2, f' {target_label}/{attr_label}(実数)', font=_SUBTITLE_FONT)
    row += 1
    row = _write_target_group_header(ws, row, target_label, n_opts, first_col)
    _write_index_and_pct_labels(ws, row, target_cols, first_col, total_col, index_label='実数', total_label=_TOTAL_LABEL)
    row += 2
    header_row = row
    for cat in [*attr_rows, _TOTAL_LABEL]:
        is_total = cat == _TOTAL_LABEL
        _grid(ws, row, 4, cat, align=_LABEL_ALIGN, is_total=is_total)
        for i, col in enumerate(target_cols):
            _grid(ws, row, first_col + i, int(n_df.loc[cat, col]), is_total=is_total)
        _grid(ws, row, total_col, int(n_df.loc[cat, _TOTAL_LABEL]), is_total=True)
        row += 1
    _write_attr_axis_label(ws, header_row, row - 1, attr_label)
    row += 1

    row = _write_commentary(ws, row, r)
    return row


def _write_target_group_header(ws: Worksheet, row: int, target_label: str, n_opts: int, first_col: int) -> int:
    _grid(ws, row, first_col, target_label, align=_LABEL_ALIGN)
    if n_opts > 1:
        ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=first_col + n_opts - 1)
    return row + 1


def _write_index_and_pct_labels(ws: Worksheet, row: int, target_cols: list[str], first_col: int, total_col: int,
                                 index_label: str, total_label: str) -> None:
    for i, _ in enumerate(target_cols, 1):
        _grid(ws, row, first_col + i - 1, i, align=_LABEL_ALIGN)
    _grid(ws, row, total_col, index_label, align=_LABEL_ALIGN, is_total=True)
    for i, col in enumerate(target_cols):
        _grid(ws, row + 1, first_col + i, col, align=_LABEL_ALIGN)
    _grid(ws, row + 1, total_col, total_label, align=_LABEL_ALIGN, is_total=True)


def _write_attr_axis_label(ws: Worksheet, row1: int, row2: int, attr_label: str) -> None:
    """属性設問名を表の左端（C列）にまとめて1回だけ、縦方向にマージして書く（見本と同じ配置）"""
    _set(ws, row1, 3, attr_label, align=_AXIS_ALIGN)
    _merge_if_needed(ws, row1, row2, 3)


def _write_commentary(ws: Worksheet, row: int, r: dict) -> int:
    if r.get('commentary'):
        _set(ws, row, 3, r['commentary'], font=_COMMENTARY_FONT, align=_COMMENTARY_ALIGN)
        row += 1
    return row + 1


# ---------------------------------------------------------------- 旧クロス％実数別表

def _build_cross_split_old_sheet(ws: Worksheet, cross_results: list[dict]) -> None:
    _set_common_column_widths(ws)
    row = 1
    if not cross_results:
        _set(ws, row, 1, '対象のクロス集計がありません。')
        return
    for r in cross_results:
        row = _write_cross_split_old_block(ws, row, r)


def _write_cross_split_old_block(ws: Worksheet, row: int, r: dict) -> int:
    """
    現行フォーマット（_write_cross_split_block）との違い: 「全体」を列・行とも先頭に置く
    （現行はどちらも末尾）——このフォーマットに変更する前の見た目に揃えたい既存の
    レポートテンプレート向け（ユーザーとの合意事項、2026-08-22）。
    """
    attr_label, target_label = r['attr_label'], r['target_label']
    pct_df, n_df, base = r['pct'], r['n'], r['base']
    target_cols = [c for c in pct_df.columns if c != _TOTAL_LABEL]
    attr_rows = [i for i in pct_df.index if i != _TOTAL_LABEL]
    n_opts = len(target_cols)
    first_col = 5  # E列（全体列がここに来る、その右にオプション列が続く）
    opt_start_col = first_col + 1

    _set(ws, row, 1, f'{attr_label} × {target_label}', font=_TITLE_FONT)
    row += 1

    # ％セクション
    _set(ws, row, 2, f' {target_label}/{attr_label}(％)', font=_SUBTITLE_FONT)
    row += 1
    _grid(ws, row, first_col, target_label, align=_LABEL_ALIGN)
    if n_opts > 1:
        ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=first_col + n_opts)
    row += 1
    _grid(ws, row, first_col, '%', align=_LABEL_ALIGN, is_total=True)
    for i, _ in enumerate(target_cols, 1):
        _grid(ws, row, opt_start_col + i - 1, i, align=_LABEL_ALIGN)
    row += 1
    _grid(ws, row, first_col, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    for i, col in enumerate(target_cols):
        _grid(ws, row, opt_start_col + i, col, align=_LABEL_ALIGN)
    row += 1
    header_row = row
    for cat in [_TOTAL_LABEL, *attr_rows]:
        is_total = cat == _TOTAL_LABEL
        n_val = base.get(cat, 0)
        label = f'{cat}\n(n={n_val})'
        _grid(ws, row, 4, label, align=_LABEL_ALIGN, is_total=is_total)
        _grid(ws, row, first_col, round(pct_df.loc[cat, _TOTAL_LABEL] / 100, 4),
              number_format=_PCT_FORMAT, is_total=True)
        for i, col in enumerate(target_cols):
            _grid(ws, row, opt_start_col + i, round(pct_df.loc[cat, col] / 100, 4),
                  number_format=_PCT_FORMAT, is_total=is_total)
        row += 1
    _write_attr_axis_label(ws, header_row, row - 1, attr_label)
    row += 1

    # 実数セクション
    _set(ws, row, 2, f' {target_label}/{attr_label}(実数)', font=_SUBTITLE_FONT)
    row += 1
    _grid(ws, row, first_col, target_label, align=_LABEL_ALIGN)
    if n_opts > 1:
        ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=first_col + n_opts)
    row += 1
    _grid(ws, row, first_col, '実数', align=_LABEL_ALIGN, is_total=True)
    for i, _ in enumerate(target_cols, 1):
        _grid(ws, row, opt_start_col + i - 1, i, align=_LABEL_ALIGN)
    row += 1
    _grid(ws, row, first_col, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    for i, col in enumerate(target_cols):
        _grid(ws, row, opt_start_col + i, col, align=_LABEL_ALIGN)
    row += 1
    header_row = row
    for cat in [_TOTAL_LABEL, *attr_rows]:
        is_total = cat == _TOTAL_LABEL
        _grid(ws, row, 4, cat, align=_LABEL_ALIGN, is_total=is_total)
        _grid(ws, row, first_col, int(n_df.loc[cat, _TOTAL_LABEL]), is_total=True)
        for i, col in enumerate(target_cols):
            _grid(ws, row, opt_start_col + i, int(n_df.loc[cat, col]), is_total=is_total)
        row += 1
    _write_attr_axis_label(ws, header_row, row - 1, attr_label)
    row += 1

    row = _write_commentary(ws, row, r)
    return row


# ---------------------------------------------------------------- クロス％＆実数表

def _build_cross_combined_sheet(ws: Worksheet, cross_results: list[dict]) -> None:
    _set_common_column_widths(ws)
    row = 1
    if not cross_results:
        _set(ws, row, 1, '対象のクロス集計がありません。')
        return
    for r in cross_results:
        row = _write_cross_combined_block(ws, row, r)


def _write_cross_combined_block(ws: Worksheet, row: int, r: dict) -> int:
    attr_label, target_label = r['attr_label'], r['target_label']
    pct_df, n_df = r['pct'], r['n']
    target_cols = [c for c in pct_df.columns if c != _TOTAL_LABEL]
    attr_rows = [i for i in pct_df.index if i != _TOTAL_LABEL]
    n_opts = len(target_cols)
    first_col = 5  # E列
    total_col = first_col + n_opts

    _set(ws, row, 1, f'{attr_label} × {target_label}', font=_TITLE_FONT)
    row += 1
    _set(ws, row, 2, f' {target_label}/{attr_label}', font=_SUBTITLE_FONT)
    row += 1
    row = _write_target_group_header(ws, row, target_label, n_opts, first_col)
    for i, _ in enumerate(target_cols, 1):
        _grid(ws, row, first_col + i - 1, i, align=_LABEL_ALIGN)
    row += 1
    for i, col in enumerate(target_cols):
        _grid(ws, row, first_col + i, col, align=_LABEL_ALIGN)
    _grid(ws, row, total_col, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    row += 1

    header_row = row
    for cat in [*attr_rows, _TOTAL_LABEL]:
        is_total = cat == _TOTAL_LABEL
        n_row, pct_row = row, row + 1
        _grid(ws, n_row, 4, cat, align=_LABEL_ALIGN, is_total=is_total)
        for i, col in enumerate(target_cols):
            _grid(ws, n_row, first_col + i, int(n_df.loc[cat, col]), is_total=is_total)
        _grid(ws, n_row, total_col, int(n_df.loc[cat, _TOTAL_LABEL]), is_total=True)
        for i, col in enumerate(target_cols):
            _grid(ws, pct_row, first_col + i, round(pct_df.loc[cat, col] / 100, 4),
                  number_format=_PCT_FORMAT, is_total=is_total)
        _grid(ws, pct_row, total_col, round(pct_df.loc[cat, _TOTAL_LABEL] / 100, 4),
              number_format=_PCT_FORMAT, is_total=True)
        _merge_if_needed(ws, n_row, pct_row, 4)
        row += 2
    _write_attr_axis_label(ws, header_row, row - 1, attr_label)

    _set(ws, row, first_col, '上段・実数　下段・％', font=_BODY_FONT)
    row += 2

    row = _write_commentary(ws, row, r)
    return row


# ---------------------------------------------------------------- 一覧型クロス集計表

def _build_list_cross_sheet(ws: Worksheet, groups: list[dict], sort_order: str) -> None:
    """
    対象設問ごとに、指定された複数の属性設問のクロスを1つの表にまとめて並べる（見本Excel
    「一覧型クロス集計表見本.xlsx」で確認した構成）。既存の他シートと違い、属性×対象の
    ペアごとにブロックを分けず、対象設問1つにつき「全体」行を1つだけ持たせた上で各属性の
    カテゴリ別分布を縦に並べる。列（対象設問の選択肢）はsort_orderに従って並べ替え、元の
    選択肢番号（対象設問の定義順での1始まり位置）を選択肢名の上の行に添える——その他自動
    バケット化で追加された合成列には元の番号が無いため空欄にする。
    """
    ws.column_dimensions['A'].width = 12.0
    ws.column_dimensions['B'].width = 16.0
    for i in range(3, 3 + _MAX_OPTION_COLS):
        ws.column_dimensions[_col_letter(i)].width = _OPTION_COL_WIDTH

    row = 1
    if not groups:
        _set(ws, row, 1, '対象の一覧型クロス集計がありません。')
        return
    for g in groups:
        row = _write_list_cross_group(ws, row, g, sort_order)


def order_target_labels(group: dict, sort_order: str) -> list[str]:
    """
    一覧型クロス集計の対象設問の選択肢を指定の並べ順で返す（Excel出力・画面表示の両方が使う
    共通ロジック——並べ替え自体はどちらの出力でも同じ結果になるべきため、2026-08-23に
    Excel専用の_ordered_target_labelsから公開関数として切り出した。ui/tab_crosstab_result.py
    参照）。
    """
    original_labels = group['target_labels']
    if sort_order == SORT_DESC:
        return sorted(original_labels, key=lambda label: group['overall_pct'].get(label, 0), reverse=True)
    return list(original_labels)


def _ordered_target_labels(group: dict, sort_order: str) -> tuple[list[str], dict[str, int]]:
    """(表示順のラベルリスト, ラベル→元の選択肢番号（1始まり、その他バケット等は無し）) を返す"""
    original_labels = group['target_labels']
    original_index = {label: i + 1 for i, label in enumerate(original_labels)}
    return order_target_labels(group, sort_order), original_index


def _write_list_cross_group(ws: Worksheet, row: int, group: dict, sort_order: str) -> int:
    labels, original_index = _ordered_target_labels(group, sort_order)
    n_opts = len(labels)
    first_col = 3  # C列
    total_col = first_col + n_opts

    _set(ws, row, 1, group['target_label'], font=_TITLE_FONT)
    row += 2

    row = _write_list_cross_half(ws, row, group, labels, original_index, first_col, total_col,
                                  section_label='％表', is_pct=True)
    row += 1
    row = _write_list_cross_half(ws, row, group, labels, original_index, first_col, total_col,
                                  section_label='実数表', is_pct=False)
    return row + 1


def _write_list_cross_half(ws: Worksheet, row: int, group: dict, labels: list[str],
                            original_index: dict[str, int], first_col: int, total_col: int,
                            *, section_label: str, is_pct: bool) -> int:
    # 元の選択肢番号の行（見本どおり、その他バケット等の合成列は空欄）
    for i, label in enumerate(labels):
        _grid(ws, row, first_col + i, original_index.get(label), align=_LABEL_ALIGN)
    row += 1

    # 見出し行（％表/実数表、選択肢名。実数表のみ最終列に「全体」を出す——見本の非対称な仕様）
    _grid(ws, row, 2, section_label, align=_LABEL_ALIGN)
    for i, label in enumerate(labels):
        _grid(ws, row, first_col + i, label, align=_LABEL_ALIGN)
    if not is_pct:
        _grid(ws, row, total_col, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    row += 1

    # 全体行（対象設問につき1つだけ、属性設問を横断した合計）
    _grid(ws, row, 2, _TOTAL_LABEL, align=_LABEL_ALIGN, is_total=True)
    values = group['overall_pct'] if is_pct else group['overall_n']
    for i, label in enumerate(labels):
        _write_list_cross_value(ws, row, first_col + i, values.get(label, 0), is_pct=is_pct, is_total=True)
    if is_pct:
        _grid(ws, row, total_col, f"n={group['overall_base']}", align=_VALUE_ALIGN, is_total=True)
    else:
        _grid(ws, row, total_col, group['overall_base'], is_total=True)
    row += 2  # 全体行の直後は必ず1行空ける（見本どおり）

    for attr in group['attrs']:
        header_row = row
        for cat in attr['categories']:
            _grid(ws, row, 2, cat['label'], align=_LABEL_ALIGN)
            values = cat['pct'] if is_pct else cat['n']
            for i, label in enumerate(labels):
                _write_list_cross_value(ws, row, first_col + i, values.get(label, 0), is_pct=is_pct)
            if is_pct:
                _grid(ws, row, total_col, f"n={cat['base']}", align=_VALUE_ALIGN)
            else:
                _grid(ws, row, total_col, cat['base'])
            row += 1
        _set(ws, header_row, 1, attr['attr_label'], align=_AXIS_ALIGN)
        row += 1  # 属性グループの直後は1行空ける（見本どおり）

    return row


def _write_list_cross_value(ws: Worksheet, row: int, col: int, value: float, *, is_pct: bool,
                             is_total: bool = False) -> None:
    """0件/0%のセルは見本にならい空欄のままにする（罫線だけ引く）"""
    if not value:
        _grid(ws, row, col, None, is_total=is_total)
    elif is_pct:
        _grid(ws, row, col, round(value / 100, 4), number_format=_PCT_FORMAT, is_total=is_total)
    else:
        _grid(ws, row, col, int(value), is_total=is_total)
